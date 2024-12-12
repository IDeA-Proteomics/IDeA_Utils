import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import re

class SampleListFileNameException(Exception):
    pass

class HeaderNotFoundException(Exception):
    pass

class SampleNameHeaderException(Exception):
    pass

class SampleNumberException(Exception):
    pass

class SampleListReader(object):

    def __init__(self):
        self.sample_frame = None
        return

    @property
    def sample_count(self):
        return len(self.sample_frame.index)

    @property
    def samples(self):
        return sample_frame['id'].tolist()

    def readFile(self, filename, project_name = None, header_row = None, sample_name_header = None):        
        self.path = os.path.abspath(filename)
        ### if project name not provided then parse it from the sample list file name
        if project_name is None:
            split_name = self.path.split('\\')[-1]
            proj_pattern = r'([^/]+_\d{6}.*)(?=_SampleList.xlsx)'
            match = re.match(proj_pattern, split_name)
            if match:
                self.project_name = match.group(1)
            else:
                raise SampleListFileNameException()
        else:
            self.project_name = project_name

         ####  Examine XLS file.  Find row with headers and last sample (first empty after samples)
        wb = load_workbook(self.path, data_only=True)
        sh = wb.worksheets[0]

        namestr = 'sample identifier'
        numstr = 'sample number'

        self.head_row = None
        self.last_row = None

        ## Find header row by searching for 'sample number' 
        for i in range(1, sh.max_row + 1):
            if self.head_row == None:
                for j in range (1, sh.max_column + 1):
                    if sh.cell(row = i, column=j).value == namestr:
                        self.head_row = i
                        break        

        ## if no header row found, throw exception
        if self.head_row == None:
            raise HeaderNotFoundException()

        # ## make sure name header is correct or throw exception
        # if namestr not in [cell.value for cell in sh[self.head_row]]:
        #     raise SampleNameHeaderException            

        ## find last sample row (assume first empty row is end)
        for i in range(self.head_row, sh.max_row + 1):
            if all([cell.value is None for cell in sh[i]]):
                self.last_row = i
                break

        if self.last_row == None:
            self.last_row = sh.max_row + 1

        ## Read the data from the excel file with Pandas
        data = pd.read_excel(self.path, engine='openpyxl', skiprows=self.head_row - 1, nrows=(self.last_row - self.head_row))
        ## Create new data frame with only the columns we want
        self.sample_frame = pd.DataFrame(data=data[[namestr]].values, columns=['id'])

        ####  Get digits from end of sample identifier for sample number
        def parseSampleNumber(id):
            match = re.search(r'_(\d+)$', id)
            if match:
                return match.group(1)
            else:
                raise SampleNumberException()

        self.sample_frame['number'] = self.sample_frame['id'].apply(parseSampleNumber)
        self.sample_frame['method'] = "None"
        self.sample_frame['position'] = "NA"

        #########  This shouldn't be needed with the new sample lists but I'm leaving it here in case something breaks and we need it later. 
        # ## Sanitize the names to remove offensive characters
        # ### Remove offensive characters that will mess up file names or Bioinformatics code
        # def sanitizeName(name):
        #     newName = re.sub(r"[ +\[\]\.\+!/@#\$%\^&\*\(\)\?\|\\;:]+", '_', str(name))
        #     if newName[0].isdigit():
        #         newName = 'S' + newName
        #     return newName
        # self.sample_frame['name'] = self.sample_frame['name'].apply(sanitizeName)
        
        return

    ### get entry from sample table
      #  data - header name as string
      #  index - index to sample list
    def getSampleData(self, data, index):
        return self.sample_frame[data][index]


        


